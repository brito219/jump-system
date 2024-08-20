import pandas as pd
import os
import customtkinter as ctk
from tkinter import filedialog
import tkinter as tk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

def maximum(val1, val2):
    return max(val1, val2)

def format_decimal(value):
    """Formata o valor para uma casa decimal"""
    return round(value, 1)

def select(arquivos):
    dataframes = []
    print(f"Processando arquivos selecionados...")

    for path in arquivos:
        print(f"Lendo arquivo: {path}")
        df = pd.read_excel(path, header=None)

        if df.shape[0] > 9 and df.shape[1] > 13:
            nome = pd.read_excel(path, header=None).iat[2, 2]
            sexo = pd.read_excel(path, header=None).iat[2, 11]
            idade = pd.read_excel(path, header=None).iat[2, 10]
            
            dados = {
                'Nome': nome,
                'Sexo': sexo,
                'Idade': idade,
                'Potência Squat Jump': format_decimal(maximum(df.iat[7, 4], df.iat[8, 4])),
                'Altura Squat Jump': format_decimal(maximum(df.iat[7, 3], df.iat[8, 3])),
                'Potência Contramov': format_decimal(maximum(df.iat[9, 4], df.iat[10, 4])),
                'Altura Contramov': format_decimal(maximum(df.iat[9, 3], df.iat[10, 3])),
                'Potência CMJ MMSS': format_decimal(maximum(df.iat[11, 4], df.iat[12, 4])),
                'Altura CMJ MMSS': format_decimal(maximum(df.iat[11, 3], df.iat[12, 3])),
            }
            print(f"Dados extraídos: {dados}")
            dataframes.append(pd.DataFrame(dados, index=[0]))
        else:
            print(f"Arquivo {path} ignorado: formato inesperado")

    if dataframes:
        df_final = pd.concat(dataframes, ignore_index=True)
        df_final.to_excel('Banco_de_dados_GERAL.xlsx', index=False)

        # Carregar o arquivo Excel para aplicar a formatação
        wb = load_workbook('Banco_de_dados_GERAL.xlsx')
        ws = wb.active

        # Definir as cores para cada grupo de colunas
        red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

        # Definir bordas finas
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        # Aplicar as cores e bordas às colunas correspondentes
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[3].fill = red_fill   # Potência Squat Jump
            row[3].border = thin_border
            row[4].fill = red_fill   # Altura Squat Jump
            row[4].border = thin_border
            row[5].fill = green_fill # Potência Contramov
            row[5].border = thin_border
            row[6].fill = green_fill # Altura Contramov
            row[6].border = thin_border
            row[7].fill = blue_fill  # Potência CMJ MMSS
            row[7].border = thin_border
            row[8].fill = blue_fill  # Altura CMJ MMSS
            row[8].border = thin_border

        # Ajustar a largura das colunas novamente
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 5)  # Mantendo a margem extra para evitar cortes
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save('Banco_de_dados_GERAL.xlsx')

        print("Arquivo consolidado gerado com sucesso!")
    else:
        print("Nenhum dado válido encontrado para consolidar.")

def escolher_arquivos():
    arquivos = filedialog.askopenfilenames(title="Selecione os arquivos de avaliação", 
                                           filetypes=[("Arquivos Excel", "*.xlsx")])
    if arquivos:
        select(arquivos)

def centralizar(app, width, height):
    screen_width = app.winfo_screenwidth()
    screen_height = app.winfo_screenheight()
    pos_x = int((screen_width / 2) - (width / 2))
    pos_y = int((screen_height / 2) - (height / 2))
    app.geometry(f"{width}x{height}+{pos_x}+{pos_y}")

ctk.set_appearance_mode("dark") 
ctk.set_default_color_theme("dark-blue")  

app = ctk.CTk() 
centralizar(app, 800, 450)

app.title("Engenharia de dados aplicada a análise do desempenho de atletas universitários")

label = ctk.CTkLabel(app, text="Selecione as avaliações", pady=20)
label.pack()

botao = ctk.CTkButton(app, text="Escolher Arquivos", command=escolher_arquivos)
botao.pack(pady=20)

app.mainloop()
